#%% Libraries
import numpy as np
import numpy.polynomial.chebyshev as cheb
# import scipy
import pandas as pd
import scipy.optimize as scopt
import scipy.integrate as scint
# import scipy.interpolate as scintp
import matplotlib.pyplot as plt
import numba
import os
import openpyxl as xl
#%% Runs
N      = 10
trials = 14 
runs   = np.arange(trials)
path   = os.path.join(os.getcwd(), r"daData_20220629_MVPdata.xlsx")
srcfile= xl.load_workbook(path, read_only=False)

#%% Preprocessing
listdata = {}
for run in runs:
    
    # Collecting available data from experiments

    T = pd.read_excel(path, sheet_name='T')[run+1].dropna() + 273.16 # (Units = Kelvin)
    H = pd.read_excel(path, sheet_name='H')[run+1].dropna()          # (Units = %)
    V = pd.read_excel(path, sheet_name='V')[run+1].dropna()          # (Units = m/s)
    W = pd.read_excel(path, sheet_name='W')[run+1].dropna()          # (Units = grams)
    MC = pd.read_excel(path, sheet_name='MC')[run+1].dropna()        # (Units = kg/kg)
    Ngrapes  = int(srcfile['I'].cell(row = 5, column = run+3).value) # (Units = Nos)
    Wt_dym   = srcfile['I'].cell(row = 7, column = run+3).value      # (Units = grams)
    t = np.arange(len(T)) * 3600
    Wt_wat_i = W[0] - Wt_dym #kg 
    
    # Assumptions
    MW_wat   = 18.0     #g/mol
    rho_drm  = 0.001540 #g/mm3 density of raisin matrix
    rho_wat  = 0.001000 #g/mm3 density of raisin water
    vL = MW_wat/rho_wat #mm3/mol
    
    # Generating co-eff for Temperature as a function of time
    print(f"Fitting run {run+1} for T")
    def get_T(t, f, a0, a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12):
        cos = np.cos
        ft = f*t
        T = a0 + a11*t + a12*t**2
        T += a1*cos(  ft + a10) + a2*cos(2*ft+ a10) + a3*cos(3*ft+ a10)
        T += a4*cos(4*ft+ a10) + a5*cos(5*ft+ a10) + a6*cos(6*ft+ a10)
        T += a7*cos(7*ft+ a10) + a8*cos(8*ft+ a10) + a9*cos(9*ft+ a10)
        return T
    p0T = 2*np.pi*1e-5, 313, 3.0, 3.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
    pT, perrT = scopt.curve_fit(get_T, t, T, p0 = p0T)
    
    # Generating co-eff for Humidity as a function of time
    def get_RH(t, coeffs):
        RH = cheb.chebval(t, coeffs)
        return RH
    pRH = cheb.chebfit(t, H, 40)
                       
    # Dumping data in to list
    listdata[run] = {"T":pT, 
                     "RH":pRH, 
                     "Ngrapes":Ngrapes, 
                     "Mmatrix":Wt_dym, 
                     "Mwater0":Wt_wat_i, 
                     "velocity":V[0],
                     "MCsmooth":MC,
                     "t":t}

#%% Plots (Not used here)
# """
# fig = plt.figure(); 
# ax1 = fig.add_subplot(331); listdata['E01']["ax"] = ax1
# ax2 = fig.add_subplot(332); listdata['E02']["ax"] = ax2
# ax3 = fig.add_subplot(333); listdata['E03']["ax"] = ax3
# ax4 = fig.add_subplot(334); listdata['E04']["ax"] = ax4
# ax5 = fig.add_subplot(335); listdata['E05']["ax"] = ax5

# listdata['E01']['cols'] = ["A","B","C","D"]
# listdata['E02']['cols'] = ["F","G","H","I"]
# listdata['E03']['cols'] = ["K","L","M","N"]
# listdata['E04']['cols'] = ["P","Q","R","S"]
# listdata['E05']['cols'] = ["U","V","W","X"]

# path = os.path.join(os.getcwd(), "Fits.xlsx")
# try:
#     wb = xl.Workbooks(path)
# except:
#     wb = xl.Workbooks.Open(path)
    
# sheet = wb.Sheets("Plots")
# """
#%% Functions
@numba.jit(nopython = True)
def P_sat(T): #P in bar T in K
    return 10**(5.20389-1733.926/(T-39.485))

@numba.jit(nopython = True)
def get_xL(C, vL, vR):
    xL = C/(C + (1 - C*vL)/vR)
    return xL

@numba.jit(nopython = True)   
def get_lngammaL(xL, aLR, aRL, alpha):
    xR = 1-xL
    tauLR = aLR
    tauRL = aRL
    GLR = np.exp(-alpha*tauLR)
    GRL = np.exp(-alpha*tauRL)
    
    Term1 = tauRL*GRL**2/(xL + xR*GRL)**2 
    Term2 = tauLR*GLR   /(xL*GLR + xR)**2
    
    return xR**2*(Term1 + Term2)

@numba.jit(nopython = True)
def get_dlngammabydxL(xL, aLR, aRL, alpha):
    dxL = 1e-9
    lngammaL = get_lngammaL(xL, aLR, aRL, alpha)
    lngammaLplus = get_lngammaL(xL + dxL, aLR, aRL, alpha)
    return (lngammaLplus - lngammaL)/dxL

@numba.jit(nopython = True)
def get_dMbydt(dMbydt, M, Mm, NL, C, r, delr, D, vL, vR, N, aLR, aRL, alpha):
    fourpiby3 = 4*np.pi/3.0
    onethird = 1.0/3.0
    
    for i in range(N):
        V = (M[i]*vL + Mm[i]*vR)
        ri = r[i]
        C[i] = M[i]/V     
        Vip = V + fourpiby3*ri**3 
        rip = (Vip/fourpiby3)**onethird 
        dr = rip - ri
        r[i+1] = rip
        delr[i] = dr 
    
    for i in range(1,N):
        if M[i] < 0:
            print(i, M[i])
        xi = M[i]/(M[i]+Mm[i]); xim = M[i-1]/(M[i-1]+Mm[i-1])
        xborder = xim + (xi - xim)*delr[i-1]/(delr[i-1] + delr[i])
        gamma = 1 + xborder * get_dlngammabydxL(xborder, aLR, aRL, alpha)
        dCbydr = 2*(C[i] - C[i-1])/(delr[i] + delr[i-1])
        NL[i] = -D*dCbydr*gamma/(1-xborder)
    
    for i in range(N):
        dMbydt[i] = 4*np.pi*((r[i]**2)*NL[i] - (r[i+1]**2)*NL[i+1])
    
    return dMbydt
        
@numba.jit(nopython = True)
def get_MandMm(M, Mm, delr, N, C0, Cm):
    ri = 0.0
    for i in range(N):
        rip = ri + delr
        V = 4*np.pi/3*(rip**3 - ri**3)
        M[i] = V*C0
        Mm[i] = V*Cm
        
        ri = rip + 0.0
    return M, Mm
    
#%% Model: Diffusion of water from core to surface to air
def model(t, M, pRH, pT, Mm, kg, Do, TD, mwR, alpha, aLR, aRL):
    
    N = len(M)
    
    T = get_T(t, *pT)
    D = Do*np.exp(-TD/T)
    vR = mwR/rho_drm
    Psat = P_sat(T)
    RH = get_RH(t, pRH)
    Pw = RH*Psat/100

    
    xsurf = M[-1]/(M[-1] + Mm[-1])

    PL = np.exp(get_lngammaL(xsurf, aLR, aRL, alpha))*xsurf*Psat
    
    Pdrv = PL - Pw
    print(t, Pdrv)
    
    dMbydt = np.zeros(N)
    
    C = np.zeros(N)
    
    NL = np.zeros(N+1)
    NL[0]  =  0.0
    NL[N]  =  kg*Pdrv
    
    r = np.zeros(N+1)
    delr = np.zeros(N)
    
    dMbydt = get_dMbydt(dMbydt, M, Mm, NL, C, r, delr, D, vL, vR, N, aLR, aRL, alpha)

    return dMbydt

#%% Solver: Discritization, shell balance
def solve(data, kg, Do, TD, mwR, alpha, aLR, aRL):
    t = data["t"] + 0.0
    tend = t[-1]
    
    Mwater0 = data["Mwater0"]
    Mmatrix = data["Mmatrix"]
    Ngrapes = data["Ngrapes"]
    
    vR = mwR/rho_drm
    
    Q0 = Mwater0/Ngrapes/MW_wat
    Qm = Mmatrix/Ngrapes/mwR
    
    V0 = Q0*vL + Qm*vR
    R  = (3/(4*np.pi)*V0)**(1.0/3.0)
    
    delr = R/N
    
    C0 = Q0/V0
    Cm = Qm/V0
    
    Mm = np.zeros(N)
    M  = np.zeros(N)
    
    M, Mm = get_MandMm(M, Mm, delr, N, C0, Cm) 
   
    sol = scint.solve_ivp(model, [0, tend], M, 
                          args=(data["RH"], data["T"], Mm, kg, Do, TD, mwR, alpha, aLR, aRL),
                          dense_output = True, method = 'DOP853')#'LSODA')
    
    solution = sol.sol(t).transpose()

    #print(f"nfev {sol.nfev} njev {sol.njev} nlu {sol.nlu}")
    
    def get_Mwater(row):
        return np.sum(np.array(row))*MW_wat
    
    labels = [f"C{i}" for i in range(N)]
    df = pd.DataFrame(solution, columns = labels)
    df["Mwater"] = df.apply(get_Mwater, axis = 1)
    
    df_sol = pd.DataFrame({
            "t":t,
            "Ccore":solution[:,1],
            "Csurf":solution[:,-1],
            "Weight":np.array(df["Mwater"])*data["Ngrapes"] + data["Mmatrix"],
            
            })

    return df_sol

#%% Comparison Plot
"""
def plot_comparison(df_sol, data, bool_sheet = False):
    plt.pause(0.001)
    ax = data["ax"]
    ax.cla()
    t = data["df"]["t"]/3600.0
    MC = (data["df"]["Weight"] - data["Mmatrix"])/data["df"]["Weight"]
    MCsmooth = scintp.UnivariateSpline(t, MC)(t)
    MCfit = (df_sol["Weight"] - data["Mmatrix"])/df_sol["Weight"]
    ax.scatter(t, MC, s=5, alpha = 0.1, color='red')
    ax.plot(t, MCsmooth, 'g')
    ax.plot(t, MCfit, 'b')
    if bool_sheet:
        irow = 3
        n = len(t)
        cols = data["cols"]
        sheet.Range(f"{cols[0]}{irow}:{cols[0]}{irow+n-1}").Value = np.array(t).reshape((n, 1))
        sheet.Range(f"{cols[1]}{irow}:{cols[1]}{irow+n-1}").Value = np.array(MC).reshape((n,1))
        sheet.Range(f"{cols[2]}{irow}:{cols[2]}{irow+n-1}").Value = np.array(MCsmooth).reshape((n,1))
        sheet.Range(f"{cols[3]}{irow}:{cols[3]}{irow+n-1}").Value = np.array(MCfit).reshape((n,1))
"""
#%% Residual calculation
def residuals(parameters):
    print('Parameters')#,parameters) 
    parameters = [abs(params) for params in parameters]#[:-2]] + list(parameters[-2:])
    [Do, TD, mwR, alpha, aLR, aRL] = parameters[len(runs):]
    listkg = parameters[:len(runs)] 
    print("kg    :", listkg)
    print("TD    :", TD)
    print("mwR   :", mwR)
    print("alpha :", alpha)
    print("aLR   :", aLR)
    print("aRL   :", aRL)
    
    errors = []
    for irun in runs:
        kg = listkg[irun]
        data = listdata[irun]
        
        df_sol = solve(data, kg*1e-9, Do*1e-9, TD, mwR, alpha, aLR, aRL)
        MCsol = (df_sol["Weight"] - data["Mmatrix"])/df_sol["Weight"]
        error = list((MCsol - data["MCsmooth"])/data["MCsmooth"])
        
        """ 
        #Experimenal noisy data points (Not used here)
        MC = (data["df"]["Weight"] - data["Mmatrix"])/data["df"]["Weight"]
        error = list((MCsol - MC))
        """
        """
        # Comparison plots (Not used here)
        plot_comparison(df_sol, data)
        """
        errors += error
    
    # Sum of squared error
    SSE = sum(np.array(errors)**2)  
    print('SSE', SSE)   
    return errors

#%% Optimization function

# Initial Guesses
mwR   = 29.60e0
Do    = 62.00e7
TD    = 28.83e2
alpha = 17.92e-3
aLR   = 23.04e-3
aRL   = 54.90e-2

listkg = [ 8.77, 11.00, 12.01, 12.99, 14.04,
          16.00, 16.95,	14.65, 12.99, 10.99,
           9.09,  8.00,  6.50,	5.99 ]

#%%

# Solver: minimizes residuals, optimizes MTCs, Do, TD, mwR, alpha, aLR, aRL
plsq = scopt.least_squares(residuals, listkg + [Do, TD, mwR, alpha, aLR, aRL])

"""
# Old method (not used here)
plsq             = scopt.leastsq(residuals,listkg+[Do, TD, mwR, alpha, aLR, aRL],full_output = 1)
pfit             = plsq[0] # Optimized parametes
pcov             = plsq[1] # covariance
errors           = np.array(residuals(pfit))
variance         = (errors**2).sum()/(len(errors) - len(pfit))
covarianceMatrix = pcov * variance
Uncertainties    = np.diag(pcov)**0.5
"""
#%% Solution & Plots: obtained from fitted parameters, and dumped in excel sheet
for run in runs:
    kg     = listkg[run]
    df_sol = solve(listdata[run], kg*1e-9, Do*1e-9, TD, mwR, alpha, aLR, aRL)
    
    srcfile   = xl.load_workbook(path, read_only=False)   
    for i in range(len(df_sol)):
        srcfile['Sol'].cell(row = (i+2), column = run+2).value = df_sol["Weight"][i]
        srcfile['Sim'].cell(row = (i+2), column = run+2).value = df_sol["Weight"][i]/max(df_sol["Weight"])   
    srcfile['I'].cell(row = (24), column = 6).value = Do
    srcfile['I'].cell(row = (25), column = 6).value = TD
    srcfile['I'].cell(row = (26), column = 6).value = mwR
    srcfile['I'].cell(row = (27), column = 6).value = alpha
    srcfile['I'].cell(row = (28), column = 6).value = aLR
    srcfile['I'].cell(row = (29), column = 6).value = aRL    
    srcfile.save(path)
    
# Comparison Plots
Wexp = pd.read_excel(path, sheet_name='Exp')
Wsim = pd.read_excel(path, sheet_name='Sim')
t = np.linspace(0,299,300)

fits, axs = plt.subplots(ncols=7, nrows=2, figsize=(56, 16))
plt.rc('xtick', labelsize=30) 
plt.rc('ytick', labelsize=30)
for ax, i in zip(axs.flat, runs):
    
    lim = len(Wsim[i+1].dropna())
    hours = t[:lim] / t[:lim].max()
    ax.plot(hours[:lim], Wexp[i+1].dropna()[:lim], 'r-', lw = 65, alpha = 0.4)
    ax.plot(hours[:lim], Wsim[i+1].dropna()[:lim], 'k-', lw =  5)
    ax.set_xlabel('Time', fontsize = 20)
    ax.set_ylabel('Weights', fontsize = 20)

    
#%% Test Tab
 
# Returns: Optimized values of MTCs, Do, TD, mwR, alpha, aLR, aRL 
def printKg():
    listkg = list(plsq.x[:len(runs)])
    print("kg", listkg)

def printDo():
    Do,TD, mwR, alpha, aLR, aRL= plsq.x[len(runs):]
    print(plsq.x[len(runs):])
 
def tempPlot():
    plt.rc('xtick', labelsize=30) 
    plt.rc('ytick', labelsize=30)
    fits, axs = plt.subplots(ncols=4, nrows=2, figsize=(40, 20))
    for ax, i in zip(axs.flat[:7], runs[:7]):
        lim = len(Wsim[i+1].dropna())
        hours = t[:lim] / t[:lim].max()
        ax.plot(hours[:lim], Wexp[i+1].dropna()[:lim], 'b-', lw = 65, alpha = 0.3)
        ax.plot(hours[:lim], Wsim[i+1].dropna()[:lim], 'k-', lw =  5)
        
def veloPlot():
    plt.rc('xtick', labelsize=30) 
    plt.rc('ytick', labelsize=30)
    fits, axs = plt.subplots(ncols=4, nrows=2, figsize=(40, 20))
    
    for ax, i in zip(axs.flat, runs[7:]):
        lim = len(Wsim[i+1].dropna())
        hours = t[:lim] / t[:lim].max()
        ax.plot(hours[:lim], Wexp[i+1].dropna()[:lim], 'r-', lw = 65, alpha = 0.3)
        ax.plot(hours[:lim], Wsim[i+1].dropna()[:lim], 'k-', lw =  5)
        ax.set_xlabel('Time', fontsize = 30)
        ax.set_ylabel('Weights', fontsize = 30) 

#%%
# printKg()
kg =  [9.862124278842863, 11.515959569704824, 12.076475347007696, 13.98923600386242, 16.046813575526173, 25.006181094094554, 22.95017215963644, 19.64209224930133, 13.003629015809938, 12.987724856532243, 9.089685762412326, 7.999740732351474, 7.499920460122149, 6.98508645609]

# printDo()
D = [6.19959452e+08 3.38286046e+03 2.96002539e+01 1.79533545e-02
 2.30720121e-02 5.48999966e-01]
#%%
mwR      = 29.60e0  #g/mol
MW_wat   = 18.0     #g/mol

rho_drm  = 0.001540 #g/mm3 density of raisin matrix
rho_wat  = 0.001000 #g/mm3 density of raisin water

vL       = MW_wat/rho_wat #mm3/mol
vR       = mwR/rho_drm #mm3/mol
C        = df_sol["Csurf"]
X_surf   = C/(C + (1 - C*vL)/vR) #get_xL(C, vL, vR)

alpha = 17.92e-3
aLR   = 23.04e-3
aRL   = 54.90e-2
xsurf = X_surf

# run = 0
T = pd.read_excel(path, sheet_name='T')[run+1].dropna() + 273.16 # (Units = Kelvin)
H = pd.read_excel(path, sheet_name='H')[run+1].dropna() 
t     = listdata[0]["t"]

# pT    = listdata[0]["T"]
# T     = get_T(t, *pT)
Psat  = 10**(5.20389-1733.926/(T-39.485))

# def get_lngammaL(xL, aLR, aRL, alpha):
xR     = 1-xsurf
tauLR  = aLR
tauRL  = aRL
GLR    = np.exp(-alpha*tauLR)
GRL    = np.exp(-alpha*tauRL)
Term1  = tauRL*GRL**2/(xsurf + xR*GRL)**2 
Term2  = tauLR*GLR   /(xsurf*GLR + xR)**2
glngL  = xR**2*(Term1 + Term2)
P_surf = xsurf*Psat#*np.exp(glngL)*

# pH = listdata[0]["RH"]
# RH = get_RH(t, pRH)
# Pw = RH*Psat/100
P_amb = H*Psat/100

P_drv = P_surf - P_amb
import matplotlib.pyplot as plt
plt.plot(t/3600, P_drv)
plt.plot(t/3600, P_surf)
plt.plot(t/3600, xsurf)
plt.plot(t/3600, Psat)
plt.plot(t/3600, T)
plt.plot(t/3600, np.exp(glngL))
